import asyncio
import io
import math
import os
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.core.auth_middleware import CurrentUser
from app.models.estoque_minimo import EstoqueMinimo
from app.models.ordem_compra import OrdemCompraGerada, ItemOrdemCompraGerada
from app.models.parametro_global import ParametroGlobal
from app.schemas.ordem_compra import OrdemCompraGeradaResponse, GerarOCResponse, PreviewItemOC, PreviewFornecedorOC, PreviewOCResponse
from app.services import bigquery_service
from app.services.csv_service import gerar_csv_oc, calcular_quantidade_oc
from app.services.email_service import enviar_oc_por_email
from app.services.ferramentas_service import get_ferramentas
from app.services.insumos_service import get_insumos

_CONSOLIDATION_TOLERANCE_DAYS = 15


def _consolidate_dates(itens: list[PreviewItemOC]) -> list[PreviewItemOC]:
    """Agrupa itens por coorte de datas (tolerância 15 dias).
    Dentro de cada coorte usa a data máxima e marca data_entrega_ajustada=True nos que mudaram.
    Preserva a ordem original dos itens.
    """
    if len(itens) <= 1:
        return itens

    # Ordenar por data para montar coortes
    ordered = sorted(enumerate(itens), key=lambda x: x[1].data_entrega)
    coortes: list[list[int]] = []          # índices originais por coorte
    coorte_atual = [ordered[0][0]]
    coorte_inicio = date.fromisoformat(ordered[0][1].data_entrega)

    for orig_idx, item in ordered[1:]:
        item_date = date.fromisoformat(item.data_entrega)
        if (item_date - coorte_inicio).days <= _CONSOLIDATION_TOLERANCE_DAYS:
            coorte_atual.append(orig_idx)
        else:
            coortes.append(coorte_atual)
            coorte_atual = [orig_idx]
            coorte_inicio = item_date
    coortes.append(coorte_atual)

    # Aplicar data máxima por coorte
    result = list(itens)
    for coorte in coortes:
        max_date = max(date.fromisoformat(itens[i].data_entrega) for i in coorte).isoformat()
        for i in coorte:
            original = itens[i].data_entrega
            result[i] = itens[i].model_copy(update={
                "data_entrega": max_date,
                "data_entrega_ajustada": max_date != original,
            })
    return result

_TEMPLATE_PATH = Path(__file__).parent.parent.parent / "templates" / "modelo_oc.xlsx"


def _compra_sugerida(estoque_min: float, estoque_atual: float, ocs_abertas: float, moq: float) -> float:
    if estoque_min <= 0:
        return 0.0
    needed = max(0.0, estoque_min - estoque_atual - ocs_abertas)
    if needed == 0:
        return 0.0
    if moq <= 0:
        return needed
    return math.ceil(needed / moq) * moq


def _get_param(db: Session, chave: str, default: str = "") -> str:
    obj = db.query(ParametroGlobal).filter(ParametroGlobal.chave == chave).first()
    return obj.valor if obj else default


def _numero_oc(db: Session, tipo: str) -> str:
    count = db.query(OrdemCompraGerada).filter(OrdemCompraGerada.tipo == tipo).count()
    ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    return f"OC-{tipo.upper()[:3]}-{count + 1:04d}-{ts}"


async def gerar_oc_insumos(db: Session, current_user: CurrentUser) -> GerarOCResponse:
    subgrupos_raw = _get_param(db, "subgrupos_insumos", "ETIQUETA,RIBBON")
    subgrupos = [s.strip() for s in subgrupos_raw.split(",") if s.strip()]

    bq_rows = await asyncio.to_thread(bigquery_service.get_insumos, subgrupos)
    estoques: dict[str, EstoqueMinimo] = {e.cpd: e for e in db.query(EstoqueMinimo).all()}
    email_destino = _get_param(db, "email_destino_oc")

    itens_por_fornecedor: dict[str, list[dict]] = defaultdict(list)

    for row in bq_rows:
        cpd = str(row.get("CPD") or "")
        est_obj = estoques.get(cpd)
        if not est_obj:
            continue
        atual = float(row.get("ESTOQUE_ALMOXARIFADO") or 0)
        est_min = float(est_obj.estoque_minimo or 0)
        est_max = float(est_obj.estoque_maximo or 0)

        if est_min <= 0 or atual >= est_min:
            continue

        qtd = calcular_quantidade_oc(
            estoque_atual=atual,
            estoque_maximo=est_max,
            moq=float(row.get("MOQ") or 0),
            mpq=float(row.get("MPQ") or 0),
            quantidade_pendente_oc=float(row.get("QUANTIDADE_PENDENTE_OC") or 0),
        )
        if qtd <= 0:
            continue

        fornecedor = str(row.get("RAZAO_SOCIAL_FORNECEDOR") or "SEM_FORNECEDOR")
        itens_por_fornecedor[fornecedor].append({
            "cpd": cpd,
            "descricao": str(row.get("DESCRICAO_COMPLEMENTAR") or ""),
            "quantidade": qtd,
            "unidade": str(row.get("UN__MEDIDA") or ""),
            "fornecedor": fornecedor,
            "observacao": "",
            "estoque_atual": atual,
            "estoque_minimo": est_min,
        })

    if not itens_por_fornecedor:
        raise HTTPException(status_code=422, detail="Nenhum insumo abaixo do estoque mínimo.")

    ordens: list[OrdemCompraGerada] = []
    os.makedirs("oc_files", exist_ok=True)

    for fornecedor, itens in itens_por_fornecedor.items():
        filename, csv_content = gerar_csv_oc(itens, fornecedor, "insumos")
        filepath = os.path.join("oc_files", filename)

        with open(filepath, "wb") as f:
            f.write(csv_content)

        oc = OrdemCompraGerada(
            numero_oc_interno=_numero_oc(db, "insumo"),
            tipo="insumo",
            status="gerada",
            arquivo_path=filepath,
            email_destinatario=email_destino,
            gerado_por_username=current_user.username,
        )
        db.add(oc)
        db.flush()

        for item in itens:
            db.add(ItemOrdemCompraGerada(
                ordem_compra_id=oc.id,
                cpd=item["cpd"],
                descricao_item=item["descricao"],
                qtd_solicitada=item["quantidade"],
                unidade=item["unidade"],
                razao_social_fornecedor=fornecedor,
                estoque_atual=item["estoque_atual"],
                estoque_minimo_calculado=item["estoque_minimo"],
            ))

        if email_destino:
            try:
                await enviar_oc_por_email(db, email_destino, fornecedor, "insumos", filename, csv_content)
                oc.status = "enviada"
                oc.enviado_at = datetime.now(timezone.utc)
            except Exception as exc:
                oc.status = "erro"
                oc.erro_msg = str(exc)

        ordens.append(oc)

    db.commit()
    for oc in ordens:
        db.refresh(oc)

    total_itens = sum(len(itens) for itens in itens_por_fornecedor.values())
    return GerarOCResponse(
        ordens=[OrdemCompraGeradaResponse.model_validate(oc) for oc in ordens],
        total_itens=total_itens,
        mensagem=f"{len(ordens)} OC(s) gerada(s) para {total_itens} item(ns).",
    )


async def gerar_oc_ferramentas(db: Session, current_user: CurrentUser) -> GerarOCResponse:
    ferramentas = await get_ferramentas(db)
    email_destino = _get_param(db, "email_destino_oc")

    itens_por_fornecedor: dict[str, list[dict]] = defaultdict(list)

    for f in ferramentas:
        if f.situacao == "ok" or f.estoque_minimo_calculado <= 0:
            continue

        moq = f.moq if f.moq > 0 else 1
        qtd = max(f.estoque_minimo_calculado - f.estoque_atual, moq)
        qtd = math.ceil(qtd / moq) * moq if moq > 0 else qtd

        fornecedor = f.razao_social_fornecedor or "SEM_FORNECEDOR"
        itens_por_fornecedor[fornecedor].append({
            "cpd": f.cpd_ferramenta,
            "descricao": f.descricao or "",
            "quantidade": round(qtd, 2),
            "unidade": f.unidade or "",
            "fornecedor": fornecedor,
            "observacao": f"Criticidade: {f.criticidade}",
            "estoque_atual": f.estoque_atual,
            "estoque_minimo": f.estoque_minimo_calculado,
        })

    if not itens_por_fornecedor:
        raise HTTPException(status_code=422, detail="Nenhuma ferramenta abaixo do estoque mínimo calculado.")

    ordens: list[OrdemCompraGerada] = []
    os.makedirs("oc_files", exist_ok=True)

    for fornecedor, itens in itens_por_fornecedor.items():
        filename, csv_content = gerar_csv_oc(itens, fornecedor, "ferramentas")
        filepath = os.path.join("oc_files", filename)

        with open(filepath, "wb") as f_:
            f_.write(csv_content)

        oc = OrdemCompraGerada(
            numero_oc_interno=_numero_oc(db, "ferramenta"),
            tipo="ferramenta",
            status="gerada",
            arquivo_path=filepath,
            email_destinatario=email_destino,
            gerado_por_username=current_user.username,
        )
        db.add(oc)
        db.flush()

        for item in itens:
            db.add(ItemOrdemCompraGerada(
                ordem_compra_id=oc.id,
                cpd=item["cpd"],
                descricao_item=item["descricao"],
                qtd_solicitada=item["quantidade"],
                unidade=item["unidade"],
                razao_social_fornecedor=fornecedor,
                estoque_atual=item["estoque_atual"],
                estoque_minimo_calculado=item["estoque_minimo"],
            ))

        if email_destino:
            try:
                await enviar_oc_por_email(db, email_destino, fornecedor, "ferramentas", filename, csv_content)
                oc.status = "enviada"
                oc.enviado_at = datetime.now(timezone.utc)
            except Exception as exc:
                oc.status = "erro"
                oc.erro_msg = str(exc)

        ordens.append(oc)

    db.commit()
    for oc in ordens:
        db.refresh(oc)

    total_itens = sum(len(itens) for itens in itens_por_fornecedor.values())
    return GerarOCResponse(
        ordens=[OrdemCompraGeradaResponse.model_validate(oc) for oc in ordens],
        total_itens=total_itens,
        mensagem=f"{len(ordens)} OC(s) gerada(s) para {total_itens} ferramenta(s).",
    )


def _build_itens_com_entrega(items_raw, *, get_cpd, get_estoque_atual, get_estoque_min,
                              get_ocs_abertas, get_moq, get_leadtime, get_unidade,
                              get_descricao, get_codigo_fabricante, get_forn) -> dict:
    """Constrói dicionário forn→[PreviewItemOC] aplicando consolidação por coorte."""
    hoje = date.today()
    by_forn: dict[str, list[PreviewItemOC]] = defaultdict(list)

    for row in items_raw:
        qtd = _compra_sugerida(get_estoque_min(row), get_estoque_atual(row), get_ocs_abertas(row), get_moq(row))
        if qtd <= 0:
            continue
        forn = get_forn(row) or "SEM FORNECEDOR"
        lt = get_leadtime(row) or 0
        sem_lt = lt == 0
        leadtime_days = max(int(lt * 7), 1) if lt > 0 else 30
        entrega = hoje + timedelta(days=leadtime_days)
        by_forn[forn].append(PreviewItemOC(
            cpd=get_cpd(row),
            codigo_fabricante=get_codigo_fabricante(row),
            descricao=get_descricao(row),
            qtd_sugerida=qtd,
            moq=get_moq(row),
            unidade=get_unidade(row),
            leadtime_semanas=lt,
            data_entrega=entrega.isoformat(),
            sem_leadtime=sem_lt,
            estoque_atual=get_estoque_atual(row),
            estoque_minimo=get_estoque_min(row),
            ocs_abertas=get_ocs_abertas(row),
        ))

    # Consolidar datas por fornecedor
    return {forn: _consolidate_dates(itens) for forn, itens in by_forn.items()}


async def _resolver_id_map(by_forn: dict) -> dict[str, int]:
    razoes = [f for f in by_forn if f != "SEM FORNECEDOR"]
    return await asyncio.to_thread(bigquery_service.get_fornecedor_ids, razoes)


async def _build_preview_fornecedores(db: Session) -> tuple[list[PreviewFornecedorOC], dict[str, int]]:
    """Monta lista de fornecedores com seus itens e retorna o id_map."""
    insumos = await get_insumos(db)

    by_forn = _build_itens_com_entrega(
        insumos,
        get_cpd=lambda i: i.cpd,
        get_estoque_atual=lambda i: i.estoque_almoxarifado,
        get_estoque_min=lambda i: i.estoque_minimo,
        get_ocs_abertas=lambda i: i.ocs_abertas,
        get_moq=lambda i: i.moq,
        get_leadtime=lambda i: i.leadtime_semanas,
        get_unidade=lambda i: i.unidade,
        get_descricao=lambda i: i.descricao,
        get_codigo_fabricante=lambda i: i.codigo_fabricante,
        get_forn=lambda i: i.razao_social_fornecedor,
    )

    if not by_forn:
        raise HTTPException(status_code=422, detail="Nenhum insumo precisa de reposição no momento.")

    id_map = await _resolver_id_map(by_forn)
    fornecedores = [
        PreviewFornecedorOC(
            razao_social=forn,
            id_fornecedor=id_map.get(forn.strip().upper()),
            itens=itens,
        )
        for forn, itens in sorted(by_forn.items())
    ]
    return fornecedores, id_map


async def _build_preview_fornecedores_ferramentas(db: Session) -> tuple[list[PreviewFornecedorOC], dict[str, int]]:
    """Mesma lógica de insumos mas usando ferramentas."""
    ferrs = await get_ferramentas(db)

    by_forn = _build_itens_com_entrega(
        ferrs,
        get_cpd=lambda f: f.cpd_ferramenta,
        get_estoque_atual=lambda f: f.estoque_atual,
        get_estoque_min=lambda f: f.estoque_minimo_calculado,
        get_ocs_abertas=lambda f: f.ocs_abertas,
        get_moq=lambda f: f.moq,
        get_leadtime=lambda f: f.leadtime_semanas,
        get_unidade=lambda f: f.unidade,
        get_descricao=lambda f: f.descricao,
        get_codigo_fabricante=lambda f: None,
        get_forn=lambda f: f.razao_social_fornecedor,
    )

    if not by_forn:
        raise HTTPException(status_code=422, detail="Nenhuma ferramenta precisa de reposição no momento.")

    id_map = await _resolver_id_map(by_forn)
    fornecedores = [
        PreviewFornecedorOC(
            razao_social=forn,
            id_fornecedor=id_map.get(forn.strip().upper()),
            itens=itens,
        )
        for forn, itens in sorted(by_forn.items())
    ]
    return fornecedores, id_map


async def get_preview_oc_insumos(db: Session) -> PreviewOCResponse:
    fornecedores, _ = await _build_preview_fornecedores(db)
    return PreviewOCResponse(
        fornecedores=fornecedores,
        total_fornecedores=len(fornecedores),
        total_itens=sum(len(f.itens) for f in fornecedores),
    )


async def get_preview_oc_ferramentas(db: Session) -> PreviewOCResponse:
    fornecedores, _ = await _build_preview_fornecedores_ferramentas(db)
    return PreviewOCResponse(
        fornecedores=fornecedores,
        total_fornecedores=len(fornecedores),
        total_itens=sum(len(f.itens) for f in fornecedores),
    )


def _gerar_excel_bytes(fornecedores: list[PreviewFornecedorOC], tipo_col: str) -> bytes:
    """Preenche o template Delphus com os itens e retorna os bytes do xlsx."""
    wb = openpyxl.load_workbook(str(_TEMPLATE_PATH))
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).value = None

    row_idx = 2
    for forn in fornecedores:
        for item in forn.itens:
            entrega = date.fromisoformat(item.data_entrega)
            ws.cell(row=row_idx, column=1).value = forn.id_fornecedor
            ws.cell(row=row_idx, column=2).value = int(item.cpd) if item.cpd.isdigit() else item.cpd
            ws.cell(row=row_idx, column=3).value = int(item.qtd_sugerida)
            ws.cell(row=row_idx, column=4).value = entrega
            ws.cell(row=row_idx, column=4).number_format = "DD/MM/YYYY"
            ws.cell(row=row_idx, column=5).value = tipo_col
            ws.cell(row=row_idx, column=6).value = "P"
            row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def gerar_excel_oc_insumos(db: Session) -> tuple[bytes, str]:
    fornecedores, _ = await _build_preview_fornecedores(db)
    filename = f"OC_Insumos_{date.today().strftime('%Y%m%d')}.xlsx"
    return _gerar_excel_bytes(fornecedores, "INSUMO"), filename


async def gerar_excel_oc_ferramentas(db: Session) -> tuple[bytes, str]:
    fornecedores, _ = await _build_preview_fornecedores_ferramentas(db)
    filename = f"OC_Ferramentas_{date.today().strftime('%Y%m%d')}.xlsx"
    return _gerar_excel_bytes(fornecedores, "FERRAMENTA"), filename
